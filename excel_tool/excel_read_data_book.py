from openpyxl import load_workbook


# handler(work_sheet, row_idx, row_count)
def read_work_book(handler, file_path, data_range=('A', 6, 7), sheet_name='Sheet1'):
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]
    row_start = data_range[1]
    row_end = data_range[2]
    row_count = 0

    for cell in ws[data_range[0]]:
        if cell.row < row_start:
            continue
        elif cell.row > row_end:
            break

        row_count += 1
        handler(ws, cell.row, row_count)
