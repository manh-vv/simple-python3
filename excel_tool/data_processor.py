from openpyxl import load_workbook

import excel_tool.excel_mongo_tool as db
from excel_tool.create_file_from_template import create_file_by_sup_name, create_quarter_by_sup_name
from excel_tool.export_month_stats import export_month_stats
from excel_tool.export_quarter_stats import export_quarter_stats
from excel_tool.variables import group_sup_cc_m_db, map_customer_code_name_db, set_sup_name_db, template_sheet_name


# Group by sup name
def group_by_sup_name():
    pipeline = [
        {
            "$group": {
                "_id": {
                    "sup_name": "$sup_name",
                    "month": "$month",
                    "customer_code": "$customer_code"
                },
                "items": {
                    "$push": "$$CURRENT"
                },
                "sum_net_sale": {
                    "$sum": "$net_sale"
                },
                "sum_vat_sale": {
                    "$sum": "$vat_sale"
                }
            }
        },
        {
            "$out": group_sup_cc_m_db
        }
    ]

    db.excel_data.aggregate(pipeline)


# Map customer code name
def map_customer_code_name():
    pipeline = [
        {
            "$group": {
                "_id": "$customer_code",
                "customer_name": {"$first": "$customer_name"},
                "Region": {"$first": "$Region"},
                "District": {"$first": "$District"}
            }
        },
        {
            "$out": map_customer_code_name_db
        }
    ]

    db.excel_data.aggregate(pipeline)


# Set sup name
def set_sup_name():
    pipeline = [
        {
            "$group": {
                "_id": "$sup_name"
            }
        },
        {
            "$out": set_sup_name_db
        }
    ]

    db.excel_data.aggregate(pipeline)


# output statistic by month
# py -m unittest tests/data_processor_test.py
#
def export_out1(month):
    collection_sup_name = db.db[set_sup_name_db]
    collection_sup_data = db.db[group_sup_cc_m_db]

    print('Create statistic by month and by sup_name')
    print(f'There are: {collection_sup_name.count_documents({})} sup_name items')

    all_sup_name_item = collection_sup_name.find()
    for sup_name_item in all_sup_name_item:
        sup_name = sup_name_item.get('_id')

        cur_month = month.lower()
        print(f'create new file: sup_name={sup_name} month={cur_month}')
        new_file = create_file_by_sup_name('template1.xlsx', sup_name, cur_month)
        current_book = load_workbook(filename=new_file)

        # find by sup_name and month
        _filter = {
            '_id.sup_name': sup_name,
            '_id.month': month
        }
        all_month_sup_name_item = collection_sup_data.find(_filter)
        for month_sup_name_item in all_month_sup_name_item:
            export_month_stats(sup_name, cur_month, month_sup_name_item, current_book)

        current_book.remove(current_book[template_sheet_name])
        current_book.save(new_file)


# output statistic by month
# py -m unittest tests/data_processor_test.py
#
def export_incentive():
    print('Create quarter statistic by sup_name')
    collection_sup_data = db.db[group_sup_cc_m_db]

    data = collection_sup_data.aggregate([
        {
            "$addFields": {
                "sup_name": "$_id.sup_name",
                "month": "$_id.month",
                "customer_code": "$_id.customer_code",
            }
        },
        {
            "$project": {
                "_id": 0,
                "sup_name": 1,
                "month": 1,
                "customer_code": 1,
                "sum_net_sale": 1,
                "sum_vat_sale": 1
            }
        },
        {
            "$group": {
                "_id": "$sup_name",
                "items": {
                    "$push": "$$CURRENT"
                }
            }
        }
    ])

    for sup in list(data):
        sup_name = sup.get("_id")
        print(f'create new quarter file: sup_name={sup_name}')
        new_file = create_quarter_by_sup_name("template_incentive.xlsx", sup_name)
        current_book = load_workbook(filename=new_file)

        export_quarter_stats(current_book, sup.get('items'))

        current_book.remove(current_book[template_sheet_name])
        current_book.save(new_file)


def worker_data_processor():
    # export_out1('Sep')
    export_incentive()
