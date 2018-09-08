from openpyxl import load_workbook
from slugify import slugify
from shutil import copyfile

# noinspection PyUnresolvedReferences
from excel_tool_helper import month_to_num
# noinspection PyUnresolvedReferences
from read_percent import read_percent
# noinspection PyUnresolvedReferences
# from read_incentive import read_incentive

resource_folder = '../resources'
input_folder = f'{resource_folder}/input'
output_folder = f'{resource_folder}/output'

wb = load_workbook(filename=f'{input_folder}/data_book.xlsx', data_only=True)
sheet_ranges = wb['Sheet1']

# our data range is: A 6, Y 8778

# group by SUP name
sup_group = dict()
# get column Y, and fill out sup_group by sup_name=[row_num]
col_sup_name = 'X'
col_customer_code = 'C'
col_customer_name = 'D'
col_month_name = 'A'

row_data_start = 6
row_data_end = 6079

customer_code_name_mapping = dict()

row_num = 0
for sup_name_cell in sheet_ranges[col_sup_name]:
    row_num = sup_name_cell.row
    if row_num < row_data_start:
        continue

    sup_name = sup_name_cell.value

    # read customer name
    customer_code = sheet_ranges[f'{col_customer_code}{row_num}'].value
    if customer_code not in customer_code_name_mapping:
        customer_code_name_mapping[customer_code] = sheet_ranges[f'{col_customer_name}{row_num}'].value.strip()

    # read month
    cur_month = str(sheet_ranges[f'{col_month_name}{row_num}'].value).lower()

    if sup_name not in sup_group:
        sup_group[sup_name] = dict()

    month_group = sup_group[sup_name]
    if cur_month not in month_group:
        month_group[cur_month] = dict()

    customer_group = month_group[cur_month]
    if customer_code not in customer_group:
        customer_group[customer_code] = []

    customer_group[customer_code].append(row_num)

    # test
    if row_num >= row_data_end:
        break

# validate
if row_num == row_data_end:
    print('---- reach end -- finish collect and grouping data')
else:
    print('---- reading data fail: may missing row')
    raise Exception(f'reading data fail: expected {row_data_end} but get {row_num}')

# read incentive and update data in memory
# incentive_data = read_incentive(f'{input_folder}/Incentive 2017 carry FW to 2018.xlsx')
month_percent = read_percent(f'{input_folder}/phan_tram_2.xlsx')


def get_valid_sheet_name(s):
    return slugify(s)[-30:]


def get_valid_filename(s):
    return slugify(s)


def create_file_by_sup_name(sup_name, cur_month):
    file_name = get_valid_filename(sup_name)
    _file = f'{output_folder}/{cur_month}_{file_name}.xlsx'

    copyfile(f'{input_folder}/template1.xlsx', _file)
    return _file


def find_percent(_sup_name, _cur_month, _customer_code):
    if _sup_name not in month_percent:
        return 0

    if _cur_month not in month_percent[_sup_name]:
        return 0

    if _customer_code not in month_percent[_sup_name][_cur_month]:
        return 0

    return month_percent[_sup_name][_cur_month][_customer_code]


read_write_cols = [
    ('B', 'B'),  # Khu Vực
    ('F', 'C'),  # Tỉnh/Thành
    ('C', 'D'),  # Mã khách hàng
    ('D', 'E'),  # Tên khách hàng
    ('S', 'G'),  # Ngày hóa đơn
    ('Q', 'H'),  # Doanh số
    ('R', 'I')  # Thuế VAT
]

template_sheet_name = 'customer_name'
sheet_title_col = 'A1'
percent_col = 'G46'

# available rows in template now is 40
available_rows = 40
# we need to insert more rows so data can fit in

for sup_name, month_group in sup_group.items():
    print(sup_name)

    for cur_month, customer_group in month_group.items():
        print(''.ljust(5, '-'), ' ', cur_month)

        # for each sup_name create new excel file from template
        new_file = create_file_by_sup_name(sup_name, cur_month)
        if new_file == '':
            raise Exception(f'can not create file for {sup_name} -- {cur_month}')

        nwb = load_workbook(filename=new_file)

        for customer_code, row_num_list in customer_group.items():
            print(''.ljust(10, '-'), ' ', customer_code)
            # for each customer name create new sheet name from template
            work_sheet_source = nwb[template_sheet_name]
            c_sheet = nwb.copy_worksheet(work_sheet_source)
            c_sheet.title = get_valid_sheet_name(customer_code_name_mapping[customer_code])
            # fill value to sheet title
            month_num = month_to_num(cur_month)
            c_sheet[sheet_title_col] = f'Bảng kê chi tiết doanh số, chiết khấu thương mại Tháng {month_num}.2018'
            c_sheet[percent_col] = find_percent(sup_name, cur_month, customer_code)

            # then fill related values to the sheet
            # we will fill from row 5
            # need inserting more rows
            will_fill_row_count = len(row_num_list)
            if will_fill_row_count > available_rows:
                c_sheet.insert_rows(available_rows - 2, will_fill_row_count - available_rows + 5)

            row_index = 5
            for row_num in row_num_list:
                for col_name in read_write_cols:
                    c_sheet[f'{col_name[1]}{row_index}'] = sheet_ranges[f'{col_name[0]}{row_num}'].value

                row_index = row_index + 1

            print(f'---- fill out {row_index - 4}rows')

            # fill incentive data
            # if sup_name in incentive_data\
            #         and cur_month in incentive_data[sup_name]\
            #         and customer_code in incentive_data[sup_name][cur_month]:
            #
            #     for _icd in incentive_data[sup_name][cur_month][customer_code]:
            #         c_sheet[f'H{row_index}'] = _icd[0]  # doanh so
            #         c_sheet[f'I{row_index}'] = _icd[1]  # thue
            #         row_index = row_index + 1

        # remove template sheet
        nwb.remove(nwb[template_sheet_name])

        # saving
        nwb.save(new_file)
