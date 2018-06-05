from openpyxl import load_workbook
from slugify import slugify
from shutil import copyfile

wb = load_workbook(filename='input/data_book.xlsx', data_only=True)
sheet_ranges = wb['Sheet1']

# our data range is: A6, Y6749

# group by SUP name
sup_group = dict()
# get column Y, and fill out sup_group by sup_name=[row_num]
col_sup_name = 'Y'
col_customer_name = 'D'
col_month_name = 'A'

row_data_start = 6
row_data_end = 6749

row_num = 0
for sup_name_cell in sheet_ranges[col_sup_name]:
    row_num = sup_name_cell.row
    if row_num < row_data_start:
        continue

    sup_name = sup_name_cell.value

    # read customer name
    customer_name = sheet_ranges[f'{col_customer_name}{row_num}'].value.strip()

    # read month
    cur_month = sheet_ranges[f'{col_month_name}{row_num}'].value

    if sup_name in sup_group:
        cur_sup_collection = sup_group[sup_name]

        if cur_month in cur_sup_collection:
            customer_group = cur_sup_collection[cur_month]
            # check if customer name is in dictionary then add row_num
            if customer_name in customer_group:
                customer_group[customer_name].append(row_num)

            # else create new key of customer name then add row_num
            else:
                customer_group[customer_name] = [row_num]
        else:
            customer_group = cur_sup_collection[cur_month] = dict()
            customer_group[customer_name] = [row_num]
    else:
        # create customer dictionary to group customer by name
        cur_sup_collection = sup_group[sup_name] = dict()
        customer_group = cur_sup_collection[cur_month] = dict()
        customer_group[customer_name] = [row_num]

    # test
    if row_num == row_data_end:
        break

# validate
if row_num == row_data_end:
    print('---- reach end -- finish collect and grouping data')
else:
    print('---- reading data fail: may missing row')


def get_valid_sheet_name(s):
    return slugify(s)[-30:]


def get_valid_filename(s):
    return slugify(s)


def create_file_by_sup_name(sup_name, cur_month):
    file_name = get_valid_filename(sup_name)
    _file = f'output/{file_name}.xlsx'

    if cur_month == 'Jan':
        m = 1
    elif cur_month == 'Feb':
        m = 2
    elif cur_month == 'Mar':
        m = 3
    elif cur_month == 'Apr':
        m = 4
    else:
        print(''.rjust(10, '-'), 'No month num available')
        return ''

    copyfile(f'input/template{m}.xlsx', _file)
    return _file


read_write_cols = [
    ('B', 'B'),  # Khu Vực
    ('E', 'C'),  # Tỉnh/Thành
    ('C', 'D'),  # Mã khách hàng
    ('D', 'E'),  # Tên khách hàng
    ('I', 'F'),  # Số hóa đơn
    ('R', 'G'),  # Ngày hóa đơn
    ('P', 'H'),  # Doanh số
    ('S', 'I')  # Thuế VAT
]

template_sheet_name = 'customer_name'

# available rows in template now is 40
available_rows = 40
# we need to insert more rows so data can fit in

for sup_name, month_group in sup_group.items():
    print(sup_name)

    for cur_month, customer_group in month_group.items():
        print(''.ljust(5, '-'), ' ', cur_month)

        # for each sup_name create new excel file from template
        new_file = create_file_by_sup_name(sup_name, cur_month)
        if new_file == '':
            exit(1)

        nwb = load_workbook(filename=new_file)

        for customer_name, row_num_list in customer_group.items():
            print(''.ljust(10, '-'), ' ', customer_name)
            # for each customer name create new sheet name from template
            c_sheet = nwb.copy_worksheet(nwb[template_sheet_name])
            c_sheet.title = get_valid_sheet_name(customer_name)

            # then fill related values to the sheet
            # we will fill from row 5
            # need inserting more rows
            will_fill_row_count = len(row_num_list)
            if will_fill_row_count > available_rows:
                c_sheet.insert_rows(available_rows - 2, will_fill_row_count - available_rows + 2)

            row_index = 5
            for row_num in row_num_list:
                for col_name in read_write_cols:
                    c_sheet[f'{col_name[1]}{row_index}'] = sheet_ranges[f'{col_name[0]}{row_num}'].value

                row_index = row_index + 1

            print(f'---- fill out {row_index - 4}rows')

        # remove template sheet
        nwb.remove(nwb[template_sheet_name])

        # saving
        nwb.save(new_file)
